import openpyxl
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
import re


class SaveExcel(QThread):

    def __init__(self):
        super(SaveExcel, self).__init__()
        self.result_one = {}
        self.result_two = {}

    def save_excel(self, result_one, result_two):
        self.result_one = result_one
        self.result_two = result_two
        wb = openpyxl.load_workbook('maket.xlsx')
        sheet = wb.get_sheet_names()[0]
        ws = wb[sheet]
        ws['B6'] = result_one['total_received_by_account_40101_03100']
        ws['C6'] = result_two['total_received_by_account_40101_03100']
        ws['B7'] = result_one['refund_of_overpaid_amounts']
        ws['C7'] = result_two['refund_of_overpaid_amounts']
        ws['B8'] = result_one['total_transferred_to_the_budget']
        ws['C8'] = result_two['total_transferred_to_the_budget']
        ws['B9'] = result_one['consolidated_budget']
        ws['C9'] = result_two['consolidated_budget']
        ws['B10'] = result_one['article_i_federal_budget_including']
        ws['C10'] = result_two['article_i_federal_budget_including']
        ws['B11'] = result_one['vat_on_goods_sold_on_the_territory_of_the_Russian_Federation']
        ws['C11'] = result_two['vat_on_goods_sold_on_the_territory_of_the_Russian_Federation']
        ws['B12'] = result_one['vat_on_goods_imported_into_the_territory_of_the_Russian_Federation']
        ws['C12'] = result_two['vat_on_goods_imported_into_the_territory_of_the_Russian_Federation']
        ws['B13'] = result_one['income_tax']
        ws['C13'] = result_two['income_tax']
        ws['B14'] = result_one['article_II_consolidated_regional_budget']
        ws['C14'] = result_two['article_II_consolidated_regional_budget']
        ws['B16'] = result_one['regional_budgets']
        ws['C16'] = result_two['regional_budgets']
        ws['B17'] = result_one['regional_budgets_NDFL']
        ws['C17'] = result_two['regional_budgets_NDFL']
        ws['B18'] = result_one['regional_budgets_land_tax_from_organizations']
        ws['C18'] = result_two['regional_budgets_land_tax_from_organizations']
        ws['B19'] = result_one['local_budgets']
        ws['C19'] = result_two['local_budgets']
        ws['B20'] = result_one['local_budgets_NDFL']
        ws['C20'] = result_two['local_budgets_NDFL']
        ws['B21'] = result_one['local_budgets_land_tax_from_organizations']
        ws['C21'] = result_two['local_budgets_land_tax_from_organizations']
        ws['B22'] = result_one['local_budgets_comprehensive_income_taxes']
        ws['C22'] = result_two['local_budgets_comprehensive_income_taxes']
        ws['B23'] = result_one['article_III_state_off_budget_funds']
        ws['C23'] = result_two['article_III_state_off_budget_funds']
        ws['B25'] = result_one['pension_fund']
        ws['C25'] = result_two['pension_fund']
        ws['B26'] = result_one['social_insurance_fund']
        ws['C26'] = result_two['social_insurance_fund']
        ws['B27'] = result_one['federal_health_insurance_fund']
        ws['C27'] = result_two['federal_health_insurance_fund']
        ws['B28'] = result_one['territorial_health_insurance_fund']
        ws['C28'] = result_two['territorial_health_insurance_fund']
        ws['B29'] = result_one['article_IY_other_recipients_MOU_FC']
        ws['C29'] = result_two['article_IY_other_recipients_MOU_FC']
        ws['B30'] = result_one['account_balance_40101']
        ws['C30'] = result_two['account_balance_40101']
        ws['B31'] = result_one['NVS_chapter_100']
        ws['C31'] = result_two['NVS_chapter_100']
        ws['B34'] = result_one['total_for_section_III']
        ws['C34'] = result_two['total_for_section_III']
        ws['B35'] = result_one['total_for_section_III_federal_budgets']
        ws['C35'] = result_two['total_for_section_III_federal_budgets']
        ws['B36'] = result_one['total_for_section_III_regional_budgets']
        ws['C36'] = result_two['total_for_section_III_regional_budgets']
        ws['B37'] = result_one['total_for_section_III_local_budgets']
        ws['C37'] = result_two['total_for_section_III_local_budgets']
        ws['B38'] = result_one['GVF']
        ws['C38'] = result_two['GVF']

        return wb


